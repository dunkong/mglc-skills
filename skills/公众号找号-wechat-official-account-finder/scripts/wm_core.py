#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
曼格云 Skill 统一底座（所有微信生态 skill 共用，底层保持一致）

统一提供的能力：
  1. 配置读取：skill 目录 config.json → 上级 config.json → 环境变量 WM_API_KEY
  2. 无 Key 引导：标准话术 + 官方地址，退出码 3，绝不产生费用
  3. source 标识：每次调用自动带 source（body）+ X-WM-Source（header），标明来源 skill
  4. 费用预估：estimate() 零调用、零扣费，任务开始前必须调用
  5. 真实调用：call()，自动带 source、可选游标分页、统一输出标记
  6. 付费响应 24h 缓存（失败不缓存，避免重试读到旧结果）
  7. 统一退出码：0 成功 / 2 输入错误 / 3 鉴权 / 4 业务失败 / 6 网络 / 124 超时
  8. 多格式输出：Formatter 支持 json / markdown / excel / 报告 四种形态
纯标准库（Excel 需要 openpyxl，缺失时给出明确提示）。
"""
import json
import os
import re
import sys
import time
import uuid
import mimetypes
import urllib.request
import urllib.error
import urllib.parse
import hashlib

EXIT_OK, EXIT_INPUT, EXIT_AUTH, EXIT_API, EXIT_NETWORK, EXIT_TIMEOUT = 0, 2, 3, 4, 6, 124

OFFICIAL = "https://api.we-media.cn"
ROOT = "https://api.we-media.cn"

_HERE = os.path.dirname(os.path.abspath(__file__))
_META_PATH = None
for _p in (os.path.join(_HERE, "endpoints.json"),
           os.path.join(_HERE, "..", "endpoints.json"),
           os.path.join(_HERE, "..", "..", "core", "endpoints.json")):
    if os.path.isfile(_p):
        _META_PATH = _p
        break
if not _META_PATH:
    sys.stderr.write("[wm_core] 未找到 endpoints.json\n")
    sys.exit(EXIT_INPUT)

with open(_META_PATH, "r", encoding="utf-8") as _f:
    _META = json.load(_f)
EP = {e["key"]: e for e in _META["endpoints"]}
META = _META["meta"]

UPLOAD_TICKET = ROOT + "/api/v1/file-uploads/ticket"
UPLOAD_MAX_BYTES = 128 * 1024 * 1024  # 平台临时存储单文件上限 128MB

# 七牛按文件内容识别 MIME，与 mimetypes 的猜测可能不同（如 .wav 实为 audio/x-wav）。
# 申请票据时若 MIME 不符会被 403 拒绝，故这里给出实测可用的映射。
_MIME_FIX = {
    ".wav": "audio/x-wav", ".mp3": "audio/mpeg", ".m4a": "audio/mp4",
    ".aac": "audio/aac", ".ogg": "audio/ogg", ".flac": "audio/flac",
    ".mp4": "video/mp4", ".mov": "video/quicktime", ".webm": "video/webm",
    ".avi": "video/x-msvideo", ".mkv": "video/x-matroska",
}


def _guess_mime(path):
    ext = os.path.splitext(path)[1].lower()
    return _MIME_FIX.get(ext) or mimetypes.guess_type(path)[0] or "application/octet-stream"


# ---------------- 鉴权 ----------------
def load_key():
    """取 API Key：skill 目录 config.json → 上级 config.json → 环境变量。"""
    for d in (_HERE, os.path.dirname(_HERE)):
        cfg = os.path.join(d, "config.json")
        if os.path.isfile(cfg):
            try:
                v = json.load(open(cfg, "r", encoding="utf-8")).get("WM_API_KEY", "")
                if isinstance(v, str) and v.strip():
                    return v.strip()
            except Exception:
                pass
    return os.environ.get("WM_API_KEY", "").strip()


def key_url(source=None):
    """注册/创建 Key 的引导链接，带 source 统计参数。"""
    return OFFICIAL + ("?source=" + source if source else "")


KEY_GUIDE = """需要曼格云 API Key 才能继续 🔑

注册并创建 Key（约 1 分钟，全程免费）：
1. 打开 {site} 注册并登录
2. 在控制台创建 API Key（形如 ach_live_...）
3. 把 Key 发给我，我写入配置后马上开始

说明：注册、创建 Key 不产生任何费用；只有你确认执行具体查询后才按次计费。"""


def require_key(slug="", source=None):
    """无 Key 时输出标准引导并退出 3；有 Key 则返回。
    source：分发渠道标识（workbuddy/clawhub/skillhub...），写入引导链接 ?source= 供统计。"""
    k = load_key()
    if not k:
        url = key_url(source)
        sys.stderr.write(KEY_GUIDE.format(site=url) + "\n")
        print("WM_NEED_KEY=1")
        print("WM_KEY_URL=" + url)
        print("WM_OFFICIAL_SITE=" + OFFICIAL)
        if slug:
            print("WM_SKILL=" + slug)
        sys.exit(EXIT_AUTH)
    return k


# ---------------- 费用预估（零调用） ----------------
def _unit_price(ep, params):
    p = float(ep.get("price", 0.0) or 0.0)
    modes = ep.get("mode_price") or {}
    if modes:
        return float(modes.get(params.get("analysisMode", "summary"), p))
    return p


def estimate(calls, pages=1):
    """预估费用，零调用零扣费。
    calls: ["ch-info", ...] 或 [("ch-info", {...}), ...]
    pages: 每个端点的预计调用次数（分页场景传实际页数）
    返回 dict，含明细与合计。
    """
    items, total = [], 0.0
    for c in calls:
        if isinstance(c, (tuple, list)):
            key, params = c[0], (c[1] if len(c) > 1 else {})
        else:
            key, params = c, {}
        ep = EP.get(key)
        if not ep:
            items.append({"key": key, "name": "未知端点", "unit": 0.0, "times": pages, "subtotal": 0.0, "error": "unknown"})
            continue
        unit = _unit_price(ep, params)
        times = int(pages)
        sub = round(unit * times, 6)
        total += sub
        items.append({"key": key, "name": ep["name"], "unit": unit,
                      "times": times, "subtotal": sub,
                      "free": bool(ep.get("free")) or unit == 0})
    return {"items": items, "total": round(total, 6), "currency": "CNY"}


def print_estimate(est, title="费用预估"):
    """把预估结果打印成人可读 + 机器可读两种形式。"""
    print("=== {} ===".format(title))
    for it in est["items"]:
        if it.get("error"):
            print("  [!] 未知端点 {}".format(it["key"]))
            continue
        if it.get("free"):
            print("  {} — 免费".format(it["name"]))
        else:
            print("  {} — ¥{} × {} 次 = ¥{}".format(
                it["name"], it["unit"], it["times"], round(it["subtotal"], 4)))
    print("  合计约 ¥{}".format(round(est["total"], 4)))
    print("  实际以接口响应 consumption 为准。")
    print("WM_ESTIMATE_TOTAL={}".format(est["total"]))
    print("=== 预估结束 ===")


# ---------------- 调用 ----------------
class WM:
    def __init__(self, slug="mglc", timeout=90, cache=True, max_retries=2, source=None):
        self.slug = slug
        self.source = source
        self.timeout = timeout
        self.cache = cache
        self.max_retries = max_retries
        self.cache_dir = os.path.join(_HERE, ".cache")
        self.total = 0.0
        self.balance = None
        if cache:
            try:
                os.makedirs(self.cache_dir, exist_ok=True)
            except Exception:
                self.cache = False

    def _headers(self, key, method="POST"):
        h = {"X-API-Key": key, "Content-Type": "application/json"}
        h[META.get("source_header", "X-WM-Source")] = self.slug
        if method == "POST":
            h["Idempotency-Key"] = uuid.uuid4().hex
        return h

    def call(self, key, with_source=True, **params):
        """调用一个端点。返回 data。source 由 self.slug 自动注入。"""
        k = require_key(self.slug, self.source)
        ep = EP.get(key)
        if not ep:
            sys.stderr.write("未知端点 {}\n".format(key))
            sys.exit(EXIT_INPUT)

        allowed = set(ep.get("params", []))
        method = ep["method"].upper()
        base_url = ROOT + ep["path"]

        def _build(use_source):
            body = {p: v for p, v in params.items() if p in allowed and v is not None}
            if use_source:
                body[META.get("source_body_key", "source")] = self.slug
            url = base_url
            data = None
            if method == "GET":
                if body:
                    url += "?" + urllib.parse.urlencode(body)
            else:
                data = json.dumps(body, ensure_ascii=False).encode("utf-8")
            return url, data

        cache_key = None
        if self.cache and method == "POST":
            _u, _d = _build(False)
            sig = hashlib.sha256(_d or b"{}").hexdigest()[:16]
            cache_key = os.path.join(self.cache_dir, "{}_{}.json".format(key, sig))
            if os.path.isfile(cache_key):
                try:
                    cached = json.load(open(cache_key, "r", encoding="utf-8"))
                    sys.stderr.write("[缓存命中] {}（本次不再扣费）\n".format(key))
                    self._emit(cached, 0.0, cached.get("balance"), cached=True)
                    return cached.get("data", cached)
                except Exception:
                    pass

        RETRY = (429, 500, 502, 503, 504)

        def _do(use_source, depth=0):
            url, data = _build(use_source)
            for attempt in range(self.max_retries + 1):
                req = urllib.request.Request(url, data=data,
                                            headers=self._headers(k, method), method=method)
                try:
                    with urllib.request.urlopen(req, timeout=self.timeout) as r:
                        return r.read().decode("utf-8")
                except urllib.error.HTTPError as e:
                    # source 参数被网关拒绝时，自动降级重试一次
                    if use_source and e.code == 400 and depth == 0:
                        sys.stderr.write("[提示] 网关未接受 source 参数，自动降级重试。\n")
                        return _do(False, depth=1)
                    if e.code in (401, 403):
                        sys.stderr.write("鉴权失败：API Key 失效，请到 {} 重新获取或充值。\n".format(OFFICIAL))
                        sys.exit(EXIT_AUTH)
                    if e.code in RETRY and attempt < self.max_retries:
                        sys.stderr.write("[重试] HTTP {}，第{}次重试\n".format(e.code, attempt + 1))
                        time.sleep(1.5 * (attempt + 1))
                        continue
                    msg = ""
                    try:
                        msg = json.loads(e.read().decode("utf-8", "ignore")).get("message", "")
                    except Exception:
                        pass
                    sys.stderr.write("HTTP 错误 {}：{} {}\n".format(e.code, e.reason, msg))
                    sys.exit(EXIT_API)
                except (urllib.error.URLError, OSError, TimeoutError, Exception) as e:
                    if attempt < self.max_retries:
                        sys.stderr.write("[重试] 网络错误({})，第{}次重试\n".format(type(e).__name__, attempt + 1))
                        time.sleep(1.5 * (attempt + 1))
                        continue
                    sys.stderr.write("网络错误：{}\n".format(getattr(e, "reason", e)))
                    sys.exit(EXIT_NETWORK)
            sys.stderr.write("HTTP 请求失败（超过重试次数）。\n")
            sys.exit(EXIT_API)

        raw = _do(with_source)
        try:
            parsed = json.loads(raw)
        except Exception as e:
            sys.stderr.write("响应解析失败：{}\n".format(e))
            sys.exit(EXIT_API)

        code = parsed.get("code", "")
        if code and code not in ("OK", "ok") and not parsed.get("data"):
            sys.stderr.write("业务失败：code={} msg={}\n".format(
                code, parsed.get("message", parsed.get("msg", ""))))
            sys.exit(EXIT_API)

        cons = float(parsed.get("consumption", 0.0) or 0.0)
        bal = parsed.get("balance")
        if self.cache and cache_key and code in ("OK", "ok", ""):
            try:
                json.dump(parsed, open(cache_key, "w", encoding="utf-8"), ensure_ascii=False)
            except Exception:
                pass
        self._emit(parsed, cons, bal)
        return parsed.get("data", parsed)

    def paginate(self, key, max_pages=5, **params):
        """游标分页拉取：自动跟随 data.hasMore / data.cursor。
        返回合并后的列表（list of dict）。单页或非游标接口退化为一次调用。"""
        rows = []
        for _ in range(max(1, int(max_pages))):
            d = self.call(key, **params)
            items = _extract_list(d)
            if isinstance(items, list):
                rows.extend(items)
            else:
                if d is not None:
                    rows.append(d)
                break
            if isinstance(d, dict) and d.get("hasMore") and d.get("cursor"):
                params["cursor"] = d["cursor"]
            else:
                break
        return rows

    def upload_file(self, path, max_try=3):
        """把本地文件上传到平台临时存储，返回可直接喂给 AI 接口的公网 HTTPS 地址。

        对应官方「临时文件上传（给 AI Agent）」：
          1) POST /api/v1/file-uploads/ticket 申请直传票据（不计费，文件不经网关）
          2) 把 data.requiredFields 逐项作为 multipart 字段 + file 字段，POST 到 data.uploadUrl
          3) 返回 data.fileUrl，可直接作为 videoUrl / audioUrl 使用

        细节：对象存储按文件内容识别 MIME，申请票据时声明的 contentType 必须与之
        一致，否则 403。这里先按扩展名映射，若被拒则从 403 报错中取出真实 MIME 自动纠正重试。
        限制：单文件 <=128MB，票据 2 小时有效，文件约 2 小时后自动清理。
        """
        if not os.path.isfile(path):
            sys.stderr.write("文件不存在：%s\n" % path)
            sys.exit(EXIT_INPUT)
        size = os.path.getsize(path)
        if size > UPLOAD_MAX_BYTES:
            sys.stderr.write("文件超过平台存储 128MB 上限：%s（%.1fMB）\n"
                             % (path, size / 1048576.0))
            sys.exit(EXIT_INPUT)
        k = require_key(self.slug, self.source)
        fname = os.path.basename(path)
        ctype = _guess_mime(path)

        def _ticket(ct):
            body = json.dumps({"filename": fname, "bytes": size, "contentType": ct}).encode()
            req = urllib.request.Request(
                UPLOAD_TICKET, data=body, method="POST",
                headers={"X-API-Key": k, "Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.loads(r.read().decode("utf-8")).get("data") or {}

        def _push(d):
            rf = d.get("requiredFields") or {}
            b = "----wm" + uuid.uuid4().hex
            buf = b""
            for fk, fv in rf.items():
                buf += ('--%s\r\nContent-Disposition: form-data; name="%s"\r\n\r\n' % (b, fk)).encode()
                buf += str(fv).encode() + b"\r\n"
            buf += ('--%s\r\nContent-Disposition: form-data; name="file"; filename="%s"\r\n'
                    'Content-Type: %s\r\n\r\n' % (b, fname, rf.get("mimeType") or ctype)).encode()
            buf += open(path, "rb").read() + b"\r\n" + ("--%s--\r\n" % b).encode()
            req = urllib.request.Request(d["uploadUrl"], data=buf, method="POST",
                                         headers={"Content-Type": "multipart/form-data; boundary=" + b})
            try:
                with urllib.request.urlopen(req, timeout=300) as r:
                    r.read()
                return d.get("fileUrl"), None
            except urllib.error.HTTPError as e:
                return None, e.read().decode("utf-8", "ignore")

        last = ""
        for _ in range(max(1, int(max_try))):
            d = _ticket(ctype)
            if not d.get("uploadUrl"):
                sys.stderr.write("申请上传票据失败：%s\n" % json.dumps(d, ensure_ascii=False)[:200])
                sys.exit(EXIT_API)
            fu, err = _push(d)
            if fu:
                sys.stderr.write("[上传成功] %s -> %s\n" % (fname, fu))
                return fu
            last = err or ""
            m = re.search(r"this file type \(([^)]+)\) is forbidden", last)
            if m and m.group(1) != ctype:
                ctype = m.group(1)   # 按存储端识别结果纠正后重试
                continue
            break
        sys.stderr.write("上传失败：%s\n" % last[:300])
        sys.exit(EXIT_API)

    def _emit(self, parsed, cons, bal, cached=False):
        self.total += cons
        if bal is not None:
            self.balance = float(bal)
        tag = " (缓存)" if cached else ""
        sys.stderr.write("[计费]{} consumption=¥{} balance=¥{}\n".format(
            tag, cons, bal if bal is not None else "-"))
        print("WM_CONSUMPTION={}".format(cons))
        print("WM_BALANCE={}".format(bal if bal is not None else ""))

    def finish(self):
        print("WM_TOTAL_CONSUMPTION={}".format(round(self.total, 6)))
        if self.balance is not None:
            print("WM_BALANCE={}".format(self.balance))
        return self.total


def _extract_list(data):
    """从返回结构中尽力提取列表。"""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k in ("items", "accounts", "list", "records", "results",
                  "videos", "posts", "data", "platforms"):
            v = data.get(k)
            if isinstance(v, list):
                return v
            # platforms 里嵌套 items
            if isinstance(v, list) and v and isinstance(v[0], dict) and "items" in v[0]:
                out = []
                for p in v:
                    out.extend(p.get("items", []))
                return out
    return []


# ---------------- 多格式输出 ----------------
class Formatter:
    """把接口返回结构化成 json / markdown / excel / 报告 四种形态。"""
    MAX_COLS = 10
    TRUNC = 260

    def __init__(self, slug, out_dir=None):
        self.slug = slug
        if out_dir is None:
            # 默认写到 skill 根目录（scripts 的上一级）
            out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.out_dir = out_dir
        try:
            os.makedirs(self.out_dir, exist_ok=True)
        except Exception:
            self.out_dir = "."

    def _flat(self, obj):
        if not isinstance(obj, dict):
            return {"value": obj}
        out = {}
        for k, v in obj.items():
            if v is None:
                out[k] = ""
            elif isinstance(v, (dict, list)):
                s = json.dumps(v, ensure_ascii=False)
                out[k] = s if len(s) <= self.TRUNC else s[:self.TRUNC] + "…"
            else:
                out[k] = v
        return out

    def _rows(self, data):
        if isinstance(data, list):
            return [self._flat(r) for r in data]
        if isinstance(data, dict):
            rows = None
            for k in ("items", "accounts", "list", "records", "results",
                      "videos", "posts", "data"):
                if isinstance(data.get(k), list):
                    rows = data[k]
                    break
            if rows is None:
                # 尝试 platforms[].items 聚合
                pls = data.get("platforms")
                if isinstance(pls, list) and pls and isinstance(pls[0], dict) and "items" in pls[0]:
                    rows = []
                    for p in pls:
                        rows.extend(p.get("items", []))
                else:
                    rows = [data]
            return [self._flat(r) for r in rows]
        return [{"value": data}]

    def _cols(self, rows):
        seen = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.append(k)
        return seen[:self.MAX_COLS]

    def to_markdown(self, rows, title, meta=None):
        if not rows:
            return "# {}\n\n（无数据）".format(title)
        cols = self._cols(rows)
        lines = ["# " + title, ""]
        if meta:
            for mk, mv in meta.items():
                lines.append("- **{}**：{}".format(mk, mv))
            lines.append("")
        lines.append("| " + " | ".join(cols) + " |")
        lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
        for r in rows:
            cells = []
            for c in cols:
                v = r.get(c, "")
                s = str(v)
                if len(s) > self.TRUNC:
                    s = s[:self.TRUNC] + "…"
                cells.append(s.replace("\n", " ").replace("|", "/"))
            lines.append("| " + " | ".join(cells) + " |")
        return "\n".join(lines)

    def to_excel(self, rows, title, path):
        try:
            from openpyxl import Workbook
        except ImportError:
            raise RuntimeError("需要 openpyxl：pip install openpyxl（或 python -m pip install openpyxl）")
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        if rows:
            cols = self._cols(rows)
            ws.append(cols)
            for r in rows:
                ws.append([r.get(c, "") for c in cols])
            from openpyxl.styles import Font, Alignment, PatternFill
            hdr_fill = PatternFill("solid", fgColor="1F4E78")
            hdr_font = Font(bold=True, color="FFFFFF")
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # 列宽
                maxlen = max([len(str(c))] + [len(str(r.get(c, ""))) for r in rows[:200]])
                ws.column_dimensions[cell.column_letter].width = min(48, max(10, maxlen + 2))
            ws.freeze_panes = "A2"
        else:
            ws.append(["（无数据）"])
        wb.save(path)

    def _report(self, body, title):
        head = [
            "# {} · 数据报告".format(title),
            "",
            "- 生成时间：{}".format(time.strftime("%Y-%m-%d %H:%M:%S")),
            "- 数据来源：曼格云开放 API（skill：`{}`）".format(self.slug),
            "- 说明：本报告由系统自动生成，数据以官方接口实时返回为准。",
            "",
            "---",
            "",
        ]
        foot = [
            "",
            "---",
            "",
            "*本报告由曼格云数据助手自动生成。*",
        ]
        return "\n".join(head) + body + "\n".join(foot)

    def present(self, data, title, fmt="markdown", report=False, output=None, meta=None):
        """格式化并落盘。返回 (path, rows_count)。"""
        rows = self._rows(data)
        stamp = time.strftime("%Y%m%d_%H%M%S")
        fmt = (fmt or "markdown").lower()
        if fmt == "json":
            path = output or os.path.join(self.out_dir, "{}_{}.json".format(self.slug, stamp))
            json.dump(data, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        elif fmt == "excel":
            path = output or os.path.join(self.out_dir, "{}_{}.xlsx".format(self.slug, stamp))
            self.to_excel(rows, title, path)
        else:  # markdown / report
            md = self.to_markdown(rows, title, meta=meta)
            if report:
                md = self._report(md, title)
            path = output or os.path.join(self.out_dir, "{}_{}.md".format(self.slug, stamp))
            open(path, "w", encoding="utf-8").write(md)
        return path, len(rows)

    def preview(self, data, title, n=3):
        """返回前 n 行 markdown 预览（用于终端摘要）。"""
        rows = self._rows(data)[:n]
        return self.to_markdown(rows, title) if rows else "（无数据）"


# ---------------- CLI（供快速自测） ----------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: wm_core.py <estimate|list|rows> ...")
        print("\n可用端点:")
        for k, e in sorted(EP.items()):
            print("  {:32s} {:4s} ¥{}  {}".format(k, e["method"], e["price"], e["name"]))
        sys.exit(EXIT_INPUT)
    cmd = sys.argv[1]
    if cmd == "list":
        for k, e in sorted(EP.items()):
            print("{}\t{}\t{}\t{}".format(k, e["method"], e["price"], e["name"]))
    elif cmd == "estimate":
        print_estimate(estimate(sys.argv[2:]))
    elif cmd == "key":
        k = load_key()
        print("key_ready={}".format(bool(k)))
        if not k:
            print(KEY_GUIDE)
